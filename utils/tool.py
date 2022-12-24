import gzip
import time
import os
from os import path, makedirs
import openpyxl
import pandas as pd
import numpy as np
import time
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def timeStamp(timeNum):
    timeStamp = float(timeNum/1000)
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return otherStyleTime


def load_txt(load_path, first=False, sep='\t'):
    with open(load_path, 'r', encoding='utf-8', errors='ignore') as f:
        return [l.strip().split(sep)[0] if first else l.strip().split(sep) for l in f.readlines()]


def load_pattern(pattern: str, params):
    return [pattern.format(x) for x in params]


def save_txt(data: list, save_path='data.txt', sep='\t', override=False):
    dir, __ = path.split(save_path)
    if dir:
        makedirs(dir, exist_ok=True)
    with open(save_path, "w+" if override else "a+", encoding="utf-8", errors="ignore") as f:
        if data:
            for d in data:
                if d:
                    f.write(sep.join(str(item) for item in d)+'\n')


def save_xlsx(data: list, save_path='data.xlsx', table_name='Sheet', backup=False):
    dir, __ = path.split(save_path)
    if dir:
        makedirs(dir, exist_ok=True)
    try:
        read_xlsx = openpyxl.load_workbook(save_path)
    except Exception:
        read_xlsx = openpyxl.Workbook()
    if backup:
        read_xlsx.save('backup_{}.xlsx'.format(int(time.time())))
    try:
        table = read_xlsx[table_name]
    except Exception:
        read_xlsx.create_sheet(table_name)
        table = read_xlsx[table_name]
    nrows = table.max_row
    if data:
        for d in data:
            for j in range(0, len(d)):
                table.cell(nrows+1, j+1).value = ILLEGAL_CHARACTERS_RE.sub(r'',
                                                                           d[j]) if isinstance(d[j], str) else d[j]
            nrows = nrows + 1
    read_xlsx.save(save_path)


def save_raw(raw, save_path):
    dir, __ = path.split(save_path)
    if dir:
        makedirs(dir, exist_ok=True)
    with open(save_path, 'wb+') as f:
        f.write(raw)


def save_csv(data: list, save_path='data.csv', sep=',', override=False):
    dir, __ = path.split(save_path)
    if dir:
        makedirs(dir, exist_ok=True)
    pd.DataFrame(np.array(data)).to_csv(save_path, sep=sep,
                                        mode='w' if override else 'a', header=False, index=False)


def un_gz(file_path, delete_source=False):
    f_name = file_path.replace('.gz', '')
    g_file = gzip.GzipFile(file_path)
    open(f_name, "wb+").write(g_file.read())
    g_file.close()
    if delete_source:
        os.remove(file_path)
